from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from io import BytesIO

import db
import auth
from models import ReportFilters
from calc import (
    global_financials, project_financials, expense_by_category,
    income_by_category, fetch_transactions,
)
from finance import to_inr

router = APIRouter(prefix="/reports", tags=["reports"])


@router.get("")
def get_report(
    project_id: int | None = None,
    type: str | None = None,
    category: str | None = None,
    party: str | None = None,
    payment_method: str | None = None,
    from_date: str | None = None,
    to_date: str | None = None,
    user=Depends(auth.require_user),
):
    rows, count = fetch_transactions(
        project_id=project_id, type_=type, category=category, party=party,
        payment_method=payment_method, from_date=from_date, to_date=to_date,
        limit=100000,
    )
    # Period-scoped income/expense so the report matches the selected range
    period_income = sum(r["amount_paise"] for r in rows if r["type"] == "income")
    period_expense = sum(r["amount_paise"] for r in rows if r["type"] == "expense")
    if project_id:
        fin = project_financials(project_id)
        summary = {
            "scope": "Project",
            "budget_paise": fin["budget_paise"],
            "income_paise": period_income,
            "expense_paise": period_expense,
            "cash_balance_paise": period_income - period_expense,
            "remaining_budget_paise": fin["budget_paise"] - period_expense,
        }
    else:
        g = global_financials()
        summary = {
            "scope": "All Projects",
            "budget_paise": g["total_budget_paise"],
            "income_paise": period_income,
            "expense_paise": period_expense,
            "cash_balance_paise": period_income - period_expense,
            "remaining_budget_paise": g["total_budget_paise"] - period_expense,
        }
    return {
        "summary": summary,
        "expense_by_category": expense_by_category(project_id, from_date, to_date),
        "income_by_category": income_by_category(project_id, from_date, to_date),
        "transactions": [
            {k: (v / 100 if k == "amount_paise" else v) for k, v in r.items()}
            for r in rows
        ],
        "total": count,
    }


def _build_workbook(filters: ReportFilters, user):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows, count = fetch_transactions(
        project_id=filters.project_id, type_=filters.type, category=filters.category,
        party=filters.party, payment_method=filters.payment_method,
        from_date=filters.from_date, to_date=filters.to_date, limit=100000,
    )

    # Period-scoped income/expense (so report matches the selected date range)
    income = sum(r["amount_paise"] for r in rows if r["type"] == "income")
    expense = sum(r["amount_paise"] for r in rows if r["type"] == "expense")
    cash = income - expense

    if filters.project_id:
        fin = project_financials(filters.project_id)
        scope = "Project"
        budget = fin["budget_paise"]
        remain = budget - expense
        proj_name = db.get_conn().execute(
            "SELECT name FROM projects WHERE id=?", (filters.project_id,)
        ).fetchone()["name"]
    else:
        g = global_financials()
        scope = "All Projects"
        budget = g["total_budget_paise"]
        remain = budget - expense
        proj_name = "All Projects"

    period = f"{filters.from_date or '—'} to {filters.to_date or '—'}"

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1F3A5F")
    head_font = Font(bold=True, color="FFFFFF")
    money_fmt = "#,##0.00"

    # Sheet 1 — Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Platio Financial Report"])
    ws.append([f"Scope: {scope} — {proj_name}"])
    ws.append([f"Reporting Period: {period}"])
    ws.append([])
    ws.append(["Metric", "Amount (₹)"])
    for c in ("A5", "B5"):
        ws[c].fill = head_fill
        ws[c].font = head_font
    ws.append(["Budget", budget / 100])
    ws.append(["Total Income", income / 100])
    ws.append(["Total Expense", expense / 100])
    ws.append(["Current Cash Balance", cash / 100])
    ws.append(["Remaining Budget", remain / 100])
    for r in range(6, 11):
        ws[f"B{r}"].number_format = money_fmt

    # Sheet 2 — Transactions
    wt = wb.create_sheet("Transactions")
    headers = ["Date", "Transaction ID", "Project", "Type", "Category", "Subcategory",
               "Description", "Party/Vendor", "Payment Method", "Reference",
               "Amount", "Receipt Reference", "Recorded By"]
    wt.append(headers)
    for col in range(1, len(headers) + 1):
        wt.cell(row=1, column=col).fill = head_fill
        wt.cell(row=1, column=col).font = head_font
    for r in rows:
        wt.append([
            r["date"], r["id"], r["project_name"], r["type"],
            r.get("category_name") or "", r["subcategory"] or "",
            r["description"] or "", r["party"] or "", r["payment_method"] or "",
            r["reference_number"] or "", r["amount_paise"] / 100,
            ("Yes" if r.get("receipt_present") else "No"), r.get("recorded_by") or "",
        ])
    for row in range(2, len(rows) + 2):
        wt.cell(row=row, column=11).number_format = money_fmt
    wt.freeze_panes = "A2"

    # Sheet 3 — Expense Summary
    we = wb.create_sheet("Expense Summary")
    we.append(["Category", "Total Expense", "Percentage"])
    for col in range(1, 4):
        we.cell(row=1, column=col).fill = head_fill
        we.cell(row=1, column=col).font = head_font
    for e in expense_by_category(filters.project_id, filters.from_date, filters.to_date):
        we.append([e["category"], e["total_paise"] / 100, e["pct"]])
        we.cell(row=we.max_row, column=2).number_format = money_fmt
        we.cell(row=we.max_row, column=3).number_format = "0.00"

    # Sheet 4 — Income Summary
    wi = wb.create_sheet("Income Summary")
    wi.append(["Category", "Total Income", "Percentage"])
    for col in range(1, 4):
        wi.cell(row=1, column=col).fill = head_fill
        wi.cell(row=1, column=col).font = head_font
    for i in income_by_category(filters.project_id, filters.from_date, filters.to_date):
        wi.append([i["category"], i["total_paise"] / 100, i["pct"]])
        wi.cell(row=wi.max_row, column=2).number_format = money_fmt
        wi.cell(row=wi.max_row, column=3).number_format = "0.00"

    return wb


@router.post("/export")
def export_excel(filters: ReportFilters, user=Depends(auth.require_user)):
    wb = _build_workbook(filters, user)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": "attachment; filename=platio_report.xlsx"}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
