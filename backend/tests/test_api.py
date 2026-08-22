from fastapi.testclient import TestClient


def login(c: TestClient):
    r = c.post("/auth/login", json={"username": "admin", "password": "platio1234"})
    return r.status_code == 200


def create_project(c, name="Test Project", budget=1000000.0, code=None):
    body = {"name": name, "budget": budget, "status": "active"}
    if code:
        body["project_code"] = code
    r = c.post("/projects", json=body)
    assert r.status_code == 200, r.text
    return r.json()["id"]


def add_txn(c, project_id, ttype, amount, date, category=None, party=None):
    body = {"project_id": project_id, "type": ttype, "amount": amount, "date": date}
    if category:
        body["category"] = category
    if party:
        body["party"] = party
    r = c.post("/transactions", json=body)
    assert r.status_code == 200, r.text
    return r.json()["id"]


def test_login_required(c: TestClient):
    r = c.get("/projects")
    assert r.status_code == 401


def test_auth_and_me(c: TestClient):
    assert login(c)
    me = c.get("/auth/me").json()
    assert me["username"] == "admin"


def test_scenario1_financials(c: TestClient):
    login(c)
    pid = create_project(c, "Scenario1", 1000000.0)
    add_txn(c, pid, "income", 500000.0, "2026-08-01", "Client Payment")
    add_txn(c, pid, "expense", 200000.0, "2026-08-02", "Materials")
    fin = c.get(f"/projects/{pid}").json()
    assert fin["total_income"] == 500000.0
    assert fin["total_expense"] == 200000.0
    assert fin["cash_balance"] == 300000.0
    assert fin["remaining_budget"] == 800000.0
    assert fin["utilisation_pct"] == 20.0
    assert fin["budget_exceeded"] is False


def test_scenario2_isolation(c: TestClient):
    login(c)
    a = create_project(c, "ProjA", 1000000.0, code="PA")
    b = create_project(c, "ProjB", 2000000.0, code="PB")
    add_txn(c, a, "income", 500000.0, "2026-08-01")
    add_txn(c, a, "expense", 200000.0, "2026-08-02")
    add_txn(c, b, "income", 800000.0, "2026-08-03")
    add_txn(c, b, "expense", 300000.0, "2026-08-04")

    fa = c.get(f"/projects/{a}").json()
    fb = c.get(f"/projects/{b}").json()
    assert fa["total_income"] == 500000.0 and fa["total_expense"] == 200000.0
    assert fb["total_income"] == 800000.0 and fb["total_expense"] == 300000.0

    g = c.get("/dashboard").json()["global"]
    assert g["total_income_paise"] / 100 == 1300000.0
    assert g["total_expense_paise"] / 100 == 500000.0


def test_scenario6_budget_exceeded(c: TestClient):
    login(c)
    pid = create_project(c, "Over", 100000.0)  # ₹1,000 budget
    add_txn(c, pid, "expense", 150000.0, "2026-08-05", "Materials")  # ₹1,500 spend
    fin = c.get(f"/projects/{pid}").json()
    assert fin["budget_exceeded"] is True
    assert fin["remaining_budget"] == -50000.0
    # still recorded
    txns = c.get(f"/transactions?project_id={pid}").json()["items"]
    assert len(txns) == 1


def test_edit_creates_audit(c: TestClient):
    login(c)
    pid = create_project(c, "AuditProj", 500000.0)
    tid = add_txn(c, pid, "expense", 10000.0, "2026-08-06", "Labour")
    r = c.patch(f"/transactions/{tid}", json={"amount": 20000.0})
    assert r.status_code == 200
    logs = c.get("/audit-logs?entity_type=transaction").json()
    assert any(l["action"] == "transaction_edited" and l["entity_id"] == tid for l in logs)


def test_date_range_and_excel_consistency(c: TestClient):
    login(c)
    pid = create_project(c, "ReportProj", 2000000.0)
    # within range
    add_txn(c, pid, "income", 300000.0, "2026-08-10", "Client Payment")
    add_txn(c, pid, "expense", 50000.0, "2026-08-15", "Materials")
    # outside range
    add_txn(c, pid, "income", 999999.0, "2026-07-15", "Client Payment")
    add_txn(c, pid, "expense", 888888.0, "2026-09-15", "Materials")

    fd, td = "2026-08-01", "2026-08-31"
    rep = c.get(f"/reports?project_id={pid}&from_date={fd}&to_date={td}").json()
    assert rep["summary"]["income_paise"] / 100 == 300000.0
    assert rep["summary"]["expense_paise"] / 100 == 50000.0
    assert rep["total"] == 2

    # export
    r = c.post("/reports/export", json={
        "project_id": pid, "from_date": fd, "to_date": td,
    })
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]

    from io import BytesIO
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(r.content))
    assert wb.sheetnames == ["Summary", "Transactions", "Expense Summary", "Income Summary"]

    ws = wb["Summary"]
    data = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, ws.max_row + 1)}
    assert data["Total Income"] == 300000.0
    assert data["Total Expense"] == 50000.0

    wt = wb["Transactions"]
    amounts = [wt.cell(r, 11).value for r in range(2, wt.max_row + 1)]
    assert len(amounts) == 2
    assert sum(amounts) == 350000.0


def test_receipt_upload_and_serve(c: TestClient):
    login(c)
    pid = create_project(c, "ReceiptProj", 1000000.0)
    tid = add_txn(c, pid, "expense", 25000.0, "2026-08-20", "Materials")
    pdf = b"%PDF-1.1\ntest\n%%EOF"
    files = {"file": ("r.pdf", pdf, "application/pdf")}
    data = {"transaction_id": str(tid)}
    r = c.post("/receipts/upload", data=data, files=files)
    assert r.status_code == 200
    rid = r.json()["id"]
    got = c.get(f"/receipts/{rid}")
    assert got.status_code == 200
    assert got.content == pdf
